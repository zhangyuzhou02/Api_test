import yaml
import pandas
import json
from setting import DIR_NAME
from openpyxl import load_workbook

def get_yaml_data(filename):
    #读取文件
    f = open(filename,'r',encoding='utf-8')
    #获取数据
    res = yaml.load(f,Loader=yaml.FullLoader)
    return res

# def read_excel(filename,sheet_name):
#     pd = pandas.read_excel(DIR_NAME+filename,
#                            sheet_name=sheet_name,
#                            keep_default_na=False,
#                            engine='openpyxl')
#     #总行数,不包含头部
#     lines_count =  pd.shape[0]
#     #总列数,获取总列数
#     col_count = pd.columns.size
#     data = []
#     for row in range(lines_count):
#         line = []
#         for col in range(col_count):
#             text = pd.iloc[row,col]
#             if col == 1:
#                 text = json.loads(text)  # 将json格式的字符串转换成字典
#             line.append(text)
#         data.append(line)
#     return data
# def read_excel(file_name, sheet_name):
#     wb =load_workbook(file_name)
#     sheet = wb[sheet_name]
#     test_data= []
#     for i in range(2,sheet.max_row+1):
#         row_data = {}
#         row_data['测试用例名称']=sheet.cell(i,1).value
#         row_data['请求地址']=sheet.cell(i,2).value
#         row_data['请求方法']=sheet.cell(i,3).value
#         row_data['请求头部']=sheet.cell(i,4).value
#         row_data['参数']=sheet.cell(i,5).value
#         row_data['响应状态码']=sheet.cell(i,6).value
#         row_data['响应字段']=sheet.cell(i,7).value
#         test_data.append(row_data)
#     print(test_data)
#     return test_data

# 存储api的所有的信息
apis = {}
def get_api_infos(casename):
    '''
    {"api名字-登录接口":[url,method,headers,参数]}
    :return:
    '''
    res = pandas.read_excel(DIR_NAME + '/data/buyer.xlsx',
                            sheet_name=casename,
                            keep_default_na=False,
                            engine='openpyxl')
    # 获取所有的行
    row_count = res.shape[0]  # 不包括第一行  索引值从0开始

    # 获取列数
    col_count = res.columns.size # 列数
    cases_info = {}
    api_info_list = []
    for row in range(row_count):
        # api的名字
        api_name = res.iloc[row,0]
        # 定义列表 -来存储接口的信息
        for col in range(1,col_count):
            text = res.iloc[row,col]
            api_info_list.append(text)
            # 把api_name为键  api的url，method，headers和参数为元素的列表作为值
            # 存储到apis中
            apis[api_name] = api_info_list
    return apis
    # print(apis)

apis = {}
def get_cases_infos(casename):
    '''

    :param casename: 测试用例的名字-sheet页的名字
    :return:  {'casename':[
    [接口1的所有信息-url，method，headers，参数，响应提取，状态码断言，响应的断言，数据库断言，用例参数]，
    [接口2的所有的信息]
    ]
    }
    '''
    res = pandas.read_excel(DIR_NAME + '/data/buyer.xlsx',
                            sheet_name=casename,
                            keep_default_na=False,
                            engine='openpyxl')
    row_count = res.shape[0]  # 行数
    col_count = res.columns.size  # 列数
    print(col_count)
    # 字典 返回所有测试用例的信息
    cases_info = {}
    # 存储需要调用的接口的所有的信息
    steps = []
    api_infos=[]
    for row in range(row_count):
        api_name = res.iloc[row, 0]  # api的名字
        for col in range(1, col_count):
            # print(col)
            # api_infos = apis[api_name]  # 单接口的请求的相关信息的列表
            api_infos.append(res.iloc[row, col])  # api_infos包含请求和响应的所有的信息
        steps.append(api_infos)
    cases_info[casename] = steps
    print(cases_info)
    return cases_info

if __name__ == '__main__':
    get_api_infos('立即购买')