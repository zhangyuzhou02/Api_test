
import re
from openpyxl import load_workbook

class Do_Excel:
    def get_data(self, file_name, sheet_name):
        wb =load_workbook(file_name)
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['case_id'] = sheet.cell(i, 1).value
            row_data['cheack'] = sheet.cell(i, 2).value
            row_data['url'] = sheet.cell(i,3).value
            row_data['method'] = sheet.cell(i,4).value
            row_data['header'] = sheet.cell(i,5).value
            # if sheet.cell(i,5).value.find('${userId}')!=-1:
            #     userId = glv.get_value('userId')
            #     print("---------------",userId)
            #     row_data['data']=sheet.cell(i,5).value.replace('${userId}',str(userId))
            # else:
            row_data['data'] = sheet.cell(i, 6).value
            row_data['code'] = sheet.cell(i,7).value
            row_data['message'] = sheet.cell(i,8).value
            test_data.append(row_data)
        return test_data

    def write_book(self,file_name,sheet_name,i,value):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,9).value = value
        wb.save(file_name)


if __name__ == '__main__':
    print(Do_Excel().get_data('../data/buyer.xlsx', '社区'))
